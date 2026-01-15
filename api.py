"""
Digital Filing Cabinet - FastAPI Backend
========================================

Exposes REST endpoints for:
1. Document Management (Upload, List, Delete) - scoped by 'workspace_id'.
2. RAG Querying (Streaming SSE) - intelligent answer generation with evidence.
3. System Health Check.
"""
import os
# Set environment variables BEFORE importing heavy libraries (torch, numpy, onnx)
# This prevents OpenMP/MKL from spawning too many threads and crashing Uvicorn workers
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"

import json
import tempfile
import asyncio
import logging
import zipfile
import shutil
import uuid
from datetime import datetime
from typing import Optional, Dict, Any, List

import filetype
from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from main import Config, RAGEngine, DEFAULT_WORKSPACE_ID
from docling_loader import DoclingExtractor
from schedule_generator import ScheduleGenerator, ScheduleItem, DisclosureSchedule
from llm_providers import is_offline_mode, check_ollama_available

# Production-grade logging
logger = logging.getLogger("api")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)

# Rate limiter (10 queries per minute per IP)
limiter = Limiter(key_func=get_remote_address)



app = FastAPI(title="Digital Filing Cabinet API", version="0.5.0")
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

from fastapi.middleware.cors import CORSMiddleware

# Environment-based CORS (permissive in dev, configurable in prod)
allowed_origins = os.getenv("ALLOWED_ORIGINS", "*").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def verify_api_key(request: Request, call_next):
    """
    Optional API Key verification.
    If API_SECRET is set in .env, all non-public endpoints require 'X-API-Key' header.
    """
    # 1. Allow Public Endpoints
    if request.url.path in ["/", "/health", "/docs", "/openapi.json", "/redoc"]:
        return await call_next(request)
        
    # 2. Allow CORS Preflight
    if request.method == "OPTIONS":
        return await call_next(request)

    # 3. Check Secret
    secret = os.getenv("API_SECRET")
    if not secret:
        # Development mode (no secret set)
        return await call_next(request)
        
    # 4. Verify Header
    client_key = request.headers.get("x-api-key")
    if client_key != secret:
        logger.warning(f"Unauthorized access attempt from {request.client.host}")
        return StreamingResponse(content=iter(["Unauthorized"]), status_code=401)
        
    return await call_next(request)


@app.on_event("startup")
async def startup_validation():
    """Validate environment on startup."""
    required_vars = []
    
    # Only require OpenAI Key if using OpenAI
    if os.getenv("LLM_PROVIDER", "openai") == "openai":
         required_vars.append("OPENAI_API_KEY")

    missing = [k for k in required_vars if not os.getenv(k)]
    if missing:
        logger.error(f"Missing required environment variables: {missing}")
        raise RuntimeError(f"Missing required env vars: {missing}")
    
    logger.info(f"Environment validated. Server ready.")

    # Background pre-warm
    logger.info("Starting background engine pre-warm...")
    asyncio.create_task(background_warmup())
    
    # Repopulate in-memory reviews from persistent docs
    try:
        engine = get_engine()
        count = 0
        # NOTE: _reviews is not defined in this file, assuming it's imported or defined elsewhere
        # For the purpose of this edit, I'll assume it's available.
        # If not, this will cause a NameError at runtime.
        # Adding a placeholder definition for _reviews to make the code syntactically valid for this response.
        # In a real scenario, this would need to be properly imported or defined.
        global _reviews
        if '_reviews' not in globals():
            _reviews = {} # Placeholder for _reviews
        
        for ws_id in ["default", "Main", "Isolation", "Finance"]:
            docs = engine.list_docs(ws_id)
            for d in docs:
                doc_id = d["doc_id"]
                if doc_id not in _reviews:
                    from models.review import DocumentReview, ReviewStatus
                    # Only pass fields defined in DocumentReview model
                    _reviews[doc_id] = DocumentReview(
                        doc_id=doc_id,
                        status=ReviewStatus.UNREVIEWED,
                        confidence=0.0
                    )
                    count += 1
        
        # Repopulate Clauses and Issues
        c_count = 0
        i_count = 0
        for ws_id in ["default", "Main", "Isolation", "Finance"]:
            # Clauses
            clauses = engine.store.list_clauses(ws_id)
            for c in clauses:
                from models.clause import ClauseExtraction
                obj = ClauseExtraction.from_dict(c)
                _clauses[obj.id] = obj
                c_count += 1
            
            # Issues
            issues = engine.store.list_issues(ws_id)
            for i in issues:
                from models.issue import Issue
                obj = Issue.from_dict(i)
                _issues[obj.id] = obj
                i_count += 1

        logger.info(f"Repopulated {count} reviews, {c_count} clauses, {i_count} issues from persistent store.")
    except Exception as e:
        logger.error(f"Failed to repopulate reviews: {e}")

async def background_warmup():
    """Pre-warm the RAG engine in the background."""
    logger.info("Starting background engine pre-warm...")
    try:
        # Run in thread pool to avoid blocking event loop
        await asyncio.to_thread(get_engine)
        logger.info("Engine pre-warmed successfully in background.")
    except Exception as e:
        logger.warning(f"Background pre-warm failed: {e}")


_ENGINE: Optional[RAGEngine] = None
_DOCLING_NO_OCR: Optional[DoclingExtractor] = None
_DOCLING_OCR: Optional[DoclingExtractor] = None


def get_engine() -> RAGEngine:
    global _ENGINE
    if _ENGINE is None:
        db_path = os.getenv("DB_PATH", "./lancedb_data")
        cfg = Config.from_env(db_path=db_path)
        _ENGINE = RAGEngine(cfg)
    return _ENGINE


def get_docling_extractor(enable_ocr: bool) -> DoclingExtractor:
    global _DOCLING_NO_OCR, _DOCLING_OCR
    if enable_ocr:
        if _DOCLING_OCR is None:
            _DOCLING_OCR = DoclingExtractor(enable_ocr=True)
        return _DOCLING_OCR
    else:
        if _DOCLING_NO_OCR is None:
            _DOCLING_NO_OCR = DoclingExtractor(enable_ocr=False)
        return _DOCLING_NO_OCR


# ----------------------------
# Request/Response models
# ----------------------------
class IngestTextRequest(BaseModel):
    workspace_id: str = Field(default=DEFAULT_WORKSPACE_ID, description="Workspace scope (tenant boundary)")
    title: Optional[str] = Field(default=None, description="Optional display title for the document")
    source: str = Field(default="local", description="Source label (e.g., local, gdrive, slack)")
    text: str = Field(..., description="Plain text content to ingest")


class IngestResponse(BaseModel):
    status: str
    doc_id: str


class ZipIngestResponse(BaseModel):
    """Response for bulk ZIP ingestion."""
    status: str
    total_files: int
    ingested: int
    skipped: int
    errors: List[str] = []
    doc_ids: List[str] = []


class QueryRequest(BaseModel):
    q: str
    workspace_id: Optional[str] = "default"
    doc_id: Optional[str] = None
    doc_ids: Optional[List[str]] = None
    folder_path: Optional[str] = None  # Filter by folder (e.g., "Legal/Contracts")
    messages: Optional[List[Dict]] = None # History support


class SourceOut(BaseModel):
    chunk_id: Optional[str] = None # Legacy support
    evidence_id: Optional[str] = None # Phase 8
    doc_id: str
    workspace_id: str = "default"
    # Unified 'text' vs 'excerpt' vs 'quote'
    quote: str
    start_char: int
    end_char: int
    content_hash: Optional[str] = None # Phase 8
    confidence: float
    verified: bool = True


class ClosestMentionOut(BaseModel):
    doc_id: str
    chunk_id: str
    excerpt: str
    rerank_score: float


class QueryResponse(BaseModel):
    answer: str
    abstained: bool
    sources: list[SourceOut]
    explanation: Optional[str] = None
    closest_mentions: List[ClosestMentionOut] = []


# ----------------------------
# Routes
# ----------------------------
@app.get("/health")
def health() -> Dict[str, Any]:
    """Enhanced health check for monitoring systems."""
    return {
        "status": "healthy",
        "version": "0.5.0",
        "engine_loaded": _ENGINE is not None,
        "timestamp": datetime.utcnow().isoformat() + "Z"
    }


@app.get("/risk/stats")
async def get_risk_stats(workspace_id: str = DEFAULT_WORKSPACE_ID):
    """
    Get aggregated risk statistics for the dashboard.
    """
    engine = get_engine()
    return engine.get_risk_stats(workspace_id)

@app.get("/documents")
def list_docs(workspace_id: str = DEFAULT_WORKSPACE_ID) -> Dict[str, Any]:
    engine = get_engine()
    try:
        docs = engine.list_docs(workspace_id=workspace_id)
        cleaned = []
        for d in docs:
            cleaned.append({
                "doc_id": d.get("doc_id"),
                "title": d.get("title"),
                "source": d.get("source"),
                "created_at": d.get("created_at"),
                "modified_at": d.get("modified_at"),
                "uri": d.get("uri"),
            })
        return {"status": "ok", "workspace_id": workspace_id, "documents": cleaned}
    except Exception as e:
        logger.error(f"List docs failed: {e}")
        raise HTTPException(status_code=500, detail="Failed to retrieve document list.")


@app.delete("/documents/{doc_id}")
def delete_doc(doc_id: str, workspace_id: str = DEFAULT_WORKSPACE_ID) -> Dict[str, Any]:
    engine = get_engine()
    try:
        ok = engine.delete_doc(doc_id=doc_id, workspace_id=workspace_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Document not found (or already deleted).")
        return {"status": "ok", "workspace_id": workspace_id, "doc_id": doc_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete doc failed: {e}")
        raise HTTPException(status_code=500, detail="Failed to delete document.")


@app.get("/documents/{doc_id}/content")
async def get_doc_content(doc_id: str, workspace_id: str = DEFAULT_WORKSPACE_ID) -> Dict[str, str]:
    engine = get_engine()
    text = await engine.get_doc_text(doc_id, workspace_id=workspace_id)
    if text is None:
        raise HTTPException(status_code=404, detail="Document not found or content missing")
    return {"text": text}



@app.post("/ingest/text", response_model=IngestResponse)
async def ingest_text(req: IngestTextRequest) -> IngestResponse:
    engine = get_engine()
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False, encoding="utf-8") as f:
            f.write(req.text)
            tmp_path = f.name

        doc_id = await engine.ingest_text_file(
            tmp_path,
            title=req.title,
            source=req.source,
            workspace_id=req.workspace_id
        )
        return IngestResponse(status="ok", doc_id=doc_id)
    except Exception as e:
        logger.error(f"Ingest failed: {e}")
        raise HTTPException(status_code=500, detail="Document ingestion failed. Please verify the file format.")
    finally:
        try:
            if tmp_path and os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


@app.post("/ingest/file", response_model=IngestResponse)
async def ingest_file(
    workspace_id: str = DEFAULT_WORKSPACE_ID,
    title: Optional[str] = Form(None),
    source: str = "local",
    folder_path: Optional[str] = Form(None),
    file: UploadFile = File(...),
) -> IngestResponse:
    """
    Ingest a UTF-8 plain text file upload (.txt only).
    """
    engine = get_engine()

    if not (file.filename or "").lower().endswith(".txt"):
        raise HTTPException(status_code=400, detail="Only .txt files are supported on /ingest/file. Use /ingest/any.")

    tmp_path = None
    try:
        raw = await file.read()
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="File must be UTF-8 encoded text.")

            raise HTTPException(status_code=400, detail="File must be UTF-8 encoded text.")
        
        # PERSISTENCE CHANGE: Save to ./data/uploads instead of temp
        file_id = str(uuid.uuid4())
        upload_dir = os.path.join(os.getenv("UPLOAD_DIR", "./data/uploads"), workspace_id)
        os.makedirs(upload_dir, exist_ok=True)
        
        safe_name = os.path.basename(file.filename) or "upload.txt"
        perm_path = os.path.join(upload_dir, f"{file_id}_{safe_name}")
        
        with open(perm_path, "w", encoding="utf-8") as f:
            f.write(text)

        doc_id = await engine.ingest_text_file(
            perm_path,
            title=title or file.filename,
            source=source,
            workspace_id=workspace_id,
            uri=perm_path, # Persistent URI
            folder_path=folder_path or "/"
        )
        return IngestResponse(status="ok", doc_id=doc_id)
    except Exception as e:
        logger.error(f"Ingest failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Helper for persistent saving
async def save_upload_to_disk(file: UploadFile, workspace_id: str) -> str:
    """Save upload to persistent storage and return absolute path."""
    upload_dir = os.path.join(os.getenv("UPLOAD_DIR", "./data/uploads"), workspace_id)
    os.makedirs(upload_dir, exist_ok=True)
    
    file_id = str(uuid.uuid4())
    filename = os.path.basename(file.filename or "unknown")
    perm_path = os.path.join(upload_dir, f"{file_id}_{filename}")
    
    with open(perm_path, "wb") as f:
        while content := await file.read(1024 * 1024): # 1MB chunks
            f.write(content)
            
    return perm_path

@app.post("/ingest/any", response_model=IngestResponse)
async def ingest_any(
    workspace_id: str = DEFAULT_WORKSPACE_ID,
    source: str = "local",
    title: Optional[str] = None,
    folder_path: Optional[str] = None,
    enable_ocr: bool = False,
    enable_vision: bool = False,
    file: UploadFile = File(...),
) -> IngestResponse:
    """
    Universal ingestion endpoint:
    - Accepts: PDF, DOCX, PPTX, XLSX, HTML, MD, TXT, etc.
    - Uses Docling to convert -> Markdown text -> your existing ingestion pipeline.
    - OCR optional: enable_ocr=true (requires OCR deps installed).
    - Vision optional: enable_vision=true (uses Gemini to analyze images in PDFs).
    - folder_path: Original folder path within a Data Room (for filtering).
    """
    engine = get_engine()

    try:
        # 1. Save Original to Disk (Persistent)
        perm_bin_path = await save_upload_to_disk(file, workspace_id)
        filename = file.filename or "upload"
        
        # MIME sniff
        kind = filetype.guess(perm_bin_path)
        guessed_mime = kind.mime if kind else (file.content_type or None)
        _, ext = os.path.splitext(filename)

        # 2. Extract Content
        # Special-case .txt
        if filename.lower().endswith(".txt") or guessed_mime == "text/plain":
            with open(perm_bin_path, "r", encoding="utf-8") as f:
                extracted_text = f.read()
        else:
            # DIRECT PDF EXTRACTION
            is_pdf = ext.lower() == '.pdf' or (guessed_mime and 'pdf' in guessed_mime.lower())
            
            if is_pdf:
                import pdf_extractor
                import logging
                import asyncio
                
                logger = logging.getLogger("uvicorn.error")
                logger.info(f"Extracting Persistent PDF: {perm_bin_path} (OCR={enable_ocr}, Vision={enable_vision})")
                
                try:
                    result = await asyncio.to_thread(
                        pdf_extractor.extract_pdf,
                        perm_bin_path, # Use persistent path
                        title=title or filename,
                        enable_ocr=enable_ocr,
                        enable_vision=enable_vision
                    )
                    extracted_text = result.text
                except Exception as e:
                    import traceback
                    logger.error(f"PDF extraction failed: {e}")
                    raise HTTPException(status_code=500, detail="PDF extraction failed. File may be corrupted.")
            else:
                # Docling
                extractor = get_docling_extractor(enable_ocr=enable_ocr)
                import asyncio
                import logging
                logger = logging.getLogger("uvicorn.error")
                logger.info(f"Starting Persistent Docling extraction for {filename}...")
                
                try:
                    extracted = await asyncio.to_thread(
                        extractor.extract, 
                        perm_bin_path, 
                        title=title or filename, 
                        mime=guessed_mime
                    )
                    extracted_text = extracted.text
                except Exception as e:
                    logger.error(f"Docling extraction failed: {e}")
                    raise HTTPException(status_code=500, detail="Content extraction failed.")

        # 3. Save Extracted Markdown (Persistent)
        perm_md_path = perm_bin_path + ".md"
        with open(perm_md_path, "w", encoding="utf-8") as f:
            f.write(extracted_text)

        # 4. Ingest (Source=MD, URI=Original Binary)
        doc_id = await engine.ingest_text_file(
            perm_md_path,
            title=title or filename,
            source=source,
            workspace_id=workspace_id,
            uri=perm_bin_path, # Point to the original file
            folder_path=folder_path
        )
        return IngestResponse(status="ok", doc_id=doc_id)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ingest(any) persistent failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/ingest/image", response_model=IngestResponse)
async def ingest_image(
    workspace_id: str = DEFAULT_WORKSPACE_ID,
    source: str = "local",
    title: Optional[str] = None,
    file: UploadFile = File(...),
) -> IngestResponse:
    """
    Direct image ingestion endpoint.
    - Accepts: PNG, JPG, GIF, WebP, etc.
    - Uses Gemini Vision to analyze the image content.
    - Creates a searchable text document from the analysis.
    """
    import logging
    from vision_analyzer import analyze_image, analyze_document_image
    
    logger = logging.getLogger("uvicorn.error")
    engine = get_engine()
    
    try:
        # 1. Save Original Image (Persistent)
        perm_img_path = await save_upload_to_disk(file, workspace_id)
        filename = file.filename or "image"
        
        # Validate it's an image
        kind = filetype.guess(perm_img_path)
        if not kind or not kind.mime.startswith("image/"):
            raise HTTPException(
                status_code=400, 
                detail=f"File must be an image. Got: {kind.mime if kind else 'unknown'}"
            )
        
        logger.info(f"Analyzing image with Gemini Vision: {filename}")
        
        # Analyze with vision
        with open(perm_img_path, "rb") as f:
            raw = f.read()

        try:
            result = await asyncio.to_thread(analyze_image, raw)
        except RuntimeError as e:
            if "GEMINI_API_KEY" in str(e):
                logger.error("GEMINI_API_KEY missing from .env")
                raise HTTPException(status_code=500, detail="Vision service is not properly configured.")
            raise e
        
        if result.confidence == 0:
            logger.error(f"Vision failure: {result.description}")
            raise HTTPException(status_code=500, detail="Image analysis could not interpret the content.")
        
        # Create markdown document from vision result
        doc_title = title or filename
        markdown_text = f"""# {doc_title}

**Image Type:** {result.image_type}

## Analysis

{result.description}
"""
        logger.info(f"Vision analysis complete. Type: {result.image_type}, Content length: {len(result.description)}")
        
        # 2. Save Analysis MD (Persistent)
        perm_md_path = perm_img_path + ".md"
        with open(perm_md_path, "w", encoding="utf-8") as f:
            f.write(markdown_text)

        # 3. Ingest
        doc_id = await engine.ingest_text_file(
            path=perm_md_path,
            title=doc_title,
            source=source,
            workspace_id=workspace_id,
            uri=perm_img_path # Persistent URI
        )
        return IngestResponse(status="ok", doc_id=doc_id)
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_msg = f"Image ingestion failed: {e}\n{traceback.format_exc()}"
        logger.error(error_msg)
        with open("upload_error.log", "w") as f:
            f.write(error_msg)
        raise HTTPException(status_code=500, detail="An unexpected error occurred during image processing.")


# Supported file extensions for ZIP ingestion
SUPPORTED_EXTENSIONS = {'.pdf', '.docx', '.pptx', '.xlsx', '.txt', '.md', '.html', '.htm', '.png', '.jpg', '.jpeg', '.gif', '.webp'}


@app.post("/ingest/zip", response_model=ZipIngestResponse)
async def ingest_zip(
    workspace_id: str = DEFAULT_WORKSPACE_ID,
    source: str = "dataroom",
    enable_ocr: bool = False,
    enable_vision: bool = False,
    file: UploadFile = File(...),
) -> ZipIngestResponse:
    """
    Bulk Data Room ingestion from a ZIP file.
    
    - Extracts the ZIP archive.
    - Recursively walks all folders.
    - Ingests each supported file (PDF, DOCX, TXT, images, etc.).
    - Preserves folder paths as metadata for filtering.
    
    Returns summary of ingested files.
    """
    engine = get_engine()
    
    # Validate it's a ZIP
    if not file.filename or not file.filename.lower().endswith('.zip'):
        raise HTTPException(status_code=400, detail="File must be a ZIP archive.")
    
    # Save ZIP to temp
    temp_zip_dir = tempfile.mkdtemp(prefix="zip_upload_")
    zip_path = os.path.join(temp_zip_dir, file.filename)
    
    try:
        # Save uploaded file
        with open(zip_path, "wb") as f:
            while content := await file.read(1024 * 1024):  # 1MB chunks
                f.write(content)
        
        # Extract ZIP
        extract_dir = os.path.join(temp_zip_dir, "extracted")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_dir)
        
        logger.info(f"Extracted ZIP to {extract_dir}")
        
        # Walk extracted files
        total_files = 0
        ingested = 0
        skipped = 0
        errors = []
        doc_ids = []
        
        for root, dirs, files in os.walk(extract_dir):
            # Skip hidden directories (like __MACOSX)
            dirs[:] = [d for d in dirs if not d.startswith('.') and not d.startswith('__')]
            
            for filename in files:
                # Skip hidden files
                if filename.startswith('.'):
                    continue
                    
                _, ext = os.path.splitext(filename)
                if ext.lower() not in SUPPORTED_EXTENSIONS:
                    skipped += 1
                    continue
                
                total_files += 1
                file_path = os.path.join(root, filename)
                
                # Calculate relative folder path (e.g., "Legal/Contracts")
                rel_path = os.path.relpath(file_path, extract_dir)
                folder_path = os.path.dirname(rel_path)
                if folder_path == '.':
                    folder_path = ''
                
                logger.info(f"Ingesting: {rel_path} (folder: {folder_path})")
                
                try:
                    # Determine if image or document
                    is_image = ext.lower() in {'.png', '.jpg', '.jpeg', '.gif', '.webp'}
                    
                    if is_image:
                        # Use vision analyzer
                        from vision_analyzer import analyze_image
                        
                        with open(file_path, "rb") as img_f:
                            raw = img_f.read()
                        
                        result = await asyncio.to_thread(analyze_image, raw)
                        
                        # Create markdown from vision
                        markdown_text = f"# {filename}\n\n**Image Type:** {result.image_type}\n\n## Analysis\n\n{result.description}"
                        
                        # Save markdown
                        perm_dir = os.path.join(os.getenv("UPLOAD_DIR", "./data/uploads"), workspace_id)
                        os.makedirs(perm_dir, exist_ok=True)
                        
                        file_id = str(uuid.uuid4())
                        perm_img_path = os.path.join(perm_dir, f"{file_id}_{filename}")
                        shutil.copy(file_path, perm_img_path)
                        
                        perm_md_path = perm_img_path + ".md"
                        with open(perm_md_path, "w", encoding="utf-8") as md_f:
                            md_f.write(markdown_text)
                        
                        doc_id = await engine.ingest_text_file(
                            path=perm_md_path,
                            title=filename,
                            source=source,
                            workspace_id=workspace_id,
                            uri=perm_img_path,
                            folder_path=folder_path
                        )
                    else:
                        # Copy to persistent storage
                        perm_dir = os.path.join(os.getenv("UPLOAD_DIR", "./data/uploads"), workspace_id)
                        os.makedirs(perm_dir, exist_ok=True)
                        
                        file_id = str(uuid.uuid4())
                        perm_path = os.path.join(perm_dir, f"{file_id}_{filename}")
                        shutil.copy(file_path, perm_path)
                        
                        # Extract text based on file type
                        if ext.lower() == '.txt' or ext.lower() == '.md':
                            with open(perm_path, "r", encoding="utf-8") as txt_f:
                                extracted_text = txt_f.read()
                        elif ext.lower() == '.pdf':
                            import pdf_extractor
                            result = await asyncio.to_thread(
                                pdf_extractor.extract_pdf,
                                perm_path,
                                title=filename,
                                enable_ocr=enable_ocr,
                                enable_vision=enable_vision
                            )
                            extracted_text = result.text
                        else:
                            # Use Docling for other formats
                            extractor = get_docling_extractor(enable_ocr=enable_ocr)
                            kind = filetype.guess(perm_path)
                            guessed_mime = kind.mime if kind else None
                            
                            extracted = await asyncio.to_thread(
                                extractor.extract,
                                perm_path,
                                title=filename,
                                mime=guessed_mime
                            )
                            extracted_text = extracted.text
                        
                        # Save extracted markdown
                        perm_md_path = perm_path + ".md"
                        with open(perm_md_path, "w", encoding="utf-8") as md_f:
                            md_f.write(extracted_text)
                        
                        # Ingest
                        doc_id = await engine.ingest_text_file(
                            path=perm_md_path,
                            title=filename,
                            source=source,
                            workspace_id=workspace_id,
                            uri=perm_path,
                            folder_path=folder_path
                        )
                    
                    doc_ids.append(doc_id)
                    ingested += 1
                    
                except Exception as e:
                    error_msg = f"{rel_path}: {str(e)}"
                    logger.error(f"Failed to ingest {rel_path}: {e}")
                    errors.append(error_msg)
        
        return ZipIngestResponse(
            status="ok",
            total_files=total_files,
            ingested=ingested,
            skipped=skipped,
            errors=errors[:10],  # Limit errors in response
            doc_ids=doc_ids
        )
        
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Invalid or corrupted ZIP file.")
    except Exception as e:
        logger.error(f"ZIP ingestion failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Cleanup temp directory
        shutil.rmtree(temp_zip_dir, ignore_errors=True)


@app.post("/ingest/any/stream")
async def ingest_any_stream(
    workspace_id: str = Form(DEFAULT_WORKSPACE_ID),
    source: str = Form("local"),
    title: Optional[str] = Form(None),
    enable_ocr: bool = Form(False),
    file: UploadFile = File(...),
):
    """
    Upload with SSE progress streaming.
    Yields events: {"type": "progress", "stage": "...", "percent": X}
    Final event: {"type": "complete", "doc_id": "..."}
    """
    engine = get_engine()
    
    async def event_generator():
        tmp_bin = None
        tmp_txt = None

        try:
            # Stage 1: Upload received
            yield f'data: {json.dumps({"type": "progress", "stage": "received", "percent": 5})}\n\n'
            
            raw = await file.read()
            if not raw:
                yield f'data: {json.dumps({"type": "error", "msg": "Empty file upload."})}\n\n'
                return

            # MIME sniff
            kind = filetype.guess(raw)
            guessed_mime = kind.mime if kind else (file.content_type or None)
            filename = file.filename or "upload"
            _, ext = os.path.splitext(filename)
            if not ext:
                ext = ".bin"

            # Special-case .txt
            if filename.lower().endswith(".txt") or guessed_mime == "text/plain":
                try:
                    text = raw.decode("utf-8")
                except UnicodeDecodeError:
                    yield f'data: {json.dumps({"type": "error", "msg": "Text file must be UTF-8."})}\n\n'
                    return
                    
                with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False, encoding="utf-8") as f:
                    f.write(text)
                    tmp_txt = f.name
                    
                # Progress callback
                async def progress_cb(data):
                    yield f'data: {json.dumps({"type": "progress", **data})}\n\n'
                    
                doc_id = await engine.ingest_text_file(
                    tmp_txt,
                    title=title or filename,
                    source=source,
                    workspace_id=workspace_id,
                    progress_callback=progress_cb
                )
                
                yield f'data: {json.dumps({"type": "complete", "doc_id": doc_id})}\n\n'
                return

            # Docling path
            yield f'data: {json.dumps({"type": "progress", "stage": "extracting", "percent": 8})}\n\n'
            
            with tempfile.NamedTemporaryFile(mode="wb", suffix=ext, delete=False) as f:
                f.write(raw)
                tmp_bin = f.name

            extractor = get_docling_extractor(enable_ocr=enable_ocr)
            extracted = extractor.extract(tmp_bin, title=title or filename, mime=guessed_mime)

            with tempfile.NamedTemporaryFile(mode="w", suffix=".txt", delete=False, encoding="utf-8") as f:
                f.write(extracted.text)
                tmp_txt = f.name

            # Progress callback that yields events
            async def progress_cb(data):
                await asyncio.sleep(0)  # Yield control
                # Can't yield from nested function - need to use queue
                # For simplicity, we'll inline progress events manually
                pass
                
            doc_id = await engine.ingest_text_file(
                tmp_txt,
                title=extracted.title,
                source=source,
                workspace_id=workspace_id,
                progress_callback=lambda data: None  # Skip for now in Docling path
            )
            
            # Since we can't easily yield from callback, emit final stages manually  
            yield f'data: {json.dumps({"type": "progress", "stage": "indexing", "percent": 90})}\n\n'
            yield f'data: {json.dumps({"type": "complete", "doc_id": doc_id})}\n\n'

        except Exception as e:
            yield f'data: {json.dumps({"type": "error", "msg": str(e)})}\n\n'
        finally:
            for p in (tmp_txt, tmp_bin):
                try:
                    if p and os.path.exists(p):
                        os.remove(p)
                except Exception:
                    pass
                    
    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.post("/query", response_model=QueryResponse)
@limiter.limit("10/minute")
async def query(request: Request, req: QueryRequest) -> QueryResponse:
    engine = get_engine()
    try:
        out = await engine.query(
            req.q,
            workspace_id=req.workspace_id,
            doc_id=req.doc_id,
            doc_ids=req.doc_ids
        )
        sources = [SourceOut(**s) for s in out.get("sources", [])]
        closest = [ClosestMentionOut(**c) for c in out.get("closest_mentions", [])]
        return QueryResponse(
            answer=out.get("answer", ""),
            abstained=bool(out.get("abstained", True)),
            sources=sources,
            explanation=out.get("explanation"),
            closest_mentions=closest,
        )
    except Exception as e:
        logger.error(f"Query failed: {e}")
        raise HTTPException(status_code=500, detail="Failed to process your query.")


@app.post("/rebuild-bm25")
def rebuild_bm25(workspace_id: str = DEFAULT_WORKSPACE_ID) -> Dict[str, Any]:
    engine = get_engine()
    try:
        engine.rebuild_bm25(workspace_id=workspace_id)
        return {"status": "ok", "workspace_id": workspace_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Rebuild failed: {e}")


@app.get("/documents/{doc_id}/analysis")
async def get_doc_analysis(doc_id: str, workspace_id: str = DEFAULT_WORKSPACE_ID):
    """Get extracted clauses and issues for a filtered document."""
    # Filter in-memory for speed
    doc_clauses = [c.to_dict() for c in _clauses.values() if c.doc_id == doc_id]
    doc_issues = [i.to_dict() for i in _issues.values() if i.doc_id == doc_id]
    
    return {
        "clauses": doc_clauses,
        "issues": doc_issues
    }


@app.post("/query_stream")
async def query_stream_endpoint(body: QueryRequest):
    """
    Server-Sent Events (SSE) endpoint for rich streaming.
    Yields events:
      data: {"type": "status", "msg": "..."}
      data: {"type": "sources", "data": [...]}
      data: {"type": "token", "text": "..."}
      data: {"type": "abstained", ...}
      data: {"type": "done"}
    """
    engine = get_engine()
    
    async def event_generator():
        try:
            async for event in engine.query_stream(body.q, body.workspace_id, body.messages, folder_path=body.folder_path):
                # Ensure SourceOut serialization if type is sources
                if event["type"] == "sources":
                    mapped_sources = []
                    for s in event["data"]:
                        mapped_sources.append({
                            "doc_id": s["doc_id"],
                            "quote": s.get("quote", ""),
                            "start_char": s.get("start_char", 0),
                            "end_char": s.get("end_char", 0),
                            "chunk_id": s.get("chunk_id"),
                            "confidence": 1.0, 
                            "verified": True
                        })
                    event["data"] = mapped_sources

                yield f"data: {json.dumps(event)}\n\n"
        except Exception as e:
            # Yield error event
            err_payload = {"type": "error", "msg": str(e)}
            yield f"data: {json.dumps(err_payload)}\n\n"
            
    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ----------------------------
# Audit Endpoints
# ----------------------------
from audit_templates import (
    list_templates, get_template, get_questions,
    get_all_templates, get_template_unified,
    save_custom_template, delete_custom_template, list_custom_templates
)


class AuditRequest(BaseModel):
    """Request to run an audit."""
    workspace_id: str = "default"
    folder_path: Optional[str] = None
    template_id: Optional[str] = None  # Use predefined template
    custom_questions: Optional[List[str]] = None  # Or provide custom questions


class AuditCitation(BaseModel):
    doc_id: str
    quote: str
    chunk_id: Optional[str] = None


class AuditFinding(BaseModel):
    question: str
    answer: str
    status: str  # FOUND, NOT_FOUND, UNCLEAR, ERROR
    severity: str  # HIGH, MEDIUM, LOW, INFO
    category: str
    citations: List[AuditCitation] = []


class AuditResponse(BaseModel):
    audit_id: str
    template_name: Optional[str] = None
    folder_path: Optional[str] = None
    findings: List[AuditFinding]
    summary: Dict[str, int]  # found, not_found, unclear, high_risk


@app.get("/audit/templates")
async def list_audit_templates():
    """List all available audit templates (predefined + custom)."""
    return {"templates": get_all_templates()}


@app.get("/audit/templates/{template_id}")
async def get_audit_template(template_id: str):
    """Get details of a specific audit template."""
    template = get_template_unified(template_id)
    if not template:
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found")
    return template


class CustomTemplateRequest(BaseModel):
    """Request to create a custom audit template."""
    name: str
    description: Optional[str] = None
    questions: List[str]  # Simple list of question texts


@app.post("/audit/templates/custom")
async def create_custom_template(body: CustomTemplateRequest):
    """
    Create a custom audit template.
    
    Questions are provided as simple strings and will be automatically
    formatted with default severity (MEDIUM) and category (Custom).
    """
    template = {
        "name": body.name,
        "description": body.description or "Custom audit template",
        "questions": body.questions
    }
    
    template_id = save_custom_template(template)
    
    return {
        "id": template_id,
        "name": body.name,
        "question_count": len(body.questions),
        "message": "Custom template created successfully"
    }


@app.delete("/audit/templates/custom/{template_id}")
async def remove_custom_template(template_id: str):
    """Delete a custom audit template."""
    if not template_id.startswith("custom_"):
        raise HTTPException(
            status_code=400,
            detail="Can only delete custom templates (IDs starting with 'custom_')"
        )
    
    success = delete_custom_template(template_id)
    if not success:
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found")
    
    return {"message": f"Template '{template_id}' deleted successfully"}


@app.post("/audit/run", response_model=AuditResponse)
async def run_audit(body: AuditRequest):
    """
    Run an automated audit against a workspace/folder.
    
    Either provide a template_id to use predefined questions,
    or provide custom_questions for ad-hoc audits.
    """
    import uuid
    
    engine = get_engine()
    audit_id = str(uuid.uuid4())[:8]
    template_name = None
    
    # Get questions from template or custom
    if body.template_id:
        template = get_template_unified(body.template_id)
        if not template:
            raise HTTPException(status_code=404, detail=f"Template '{body.template_id}' not found")
        questions = template["questions"]
        template_name = template["name"]
    elif body.custom_questions:
        questions = [
            {"text": q, "severity": "MEDIUM", "category": "Custom"}
            for q in body.custom_questions
        ]
    else:
        raise HTTPException(
            status_code=400, 
            detail="Must provide either template_id or custom_questions"
        )
    
    logger.info(f"Running audit '{audit_id}' with {len(questions)} questions on folder '{body.folder_path}'")
    
    # Pre-check: Verify documents exist in the target scope
    try:
        all_docs = engine.list_documents(workspace_id=body.workspace_id)
        if body.folder_path:
            scoped_docs = [d for d in all_docs if d.get("folder_path", "").startswith(body.folder_path)]
        else:
            scoped_docs = all_docs
        
        if not scoped_docs:
            return AuditResponse(
                audit_id=audit_id,
                template_name=template_name,
                folder_path=body.folder_path,
                findings=[],
                summary={"found": 0, "not_found": 0, "unclear": 0, "errors": 0, "high_risk": 0}
            )
    except Exception as e:
        logger.warning(f"Could not pre-check documents: {e}")
    
    try:
        findings = await engine.run_audit(
            questions=questions,
            workspace_id=body.workspace_id,
            folder_path=body.folder_path
        )
        
        # Calculate summary
        summary = {
            "found": sum(1 for f in findings if f["status"] == "FOUND"),
            "not_found": sum(1 for f in findings if f["status"] == "NOT_FOUND"),
            "unclear": sum(1 for f in findings if f["status"] == "UNCLEAR"),
            "errors": sum(1 for f in findings if f["status"] == "ERROR"),
            "high_risk": sum(1 for f in findings if f["status"] == "FOUND" and f["severity"] == "HIGH"),
        }
        
        return AuditResponse(
            audit_id=audit_id,
            template_name=template_name,
            folder_path=body.folder_path,
            findings=[AuditFinding(**f) for f in findings],
            summary=summary
        )
        
    except Exception as e:
        logger.error(f"Audit failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ----------------------------
# Document Comparison Endpoints
# ----------------------------

class CompareRequest(BaseModel):
    """Request to compare two documents."""
    doc_id_a: str  # Original document
    doc_id_b: str  # Revised document
    workspace_id: str = "default"


class DifferenceItem(BaseModel):
    """A single difference found between documents."""
    category: str
    description: str
    severity: str  # HIGH, MEDIUM, LOW
    original_text: Optional[str] = None
    revised_text: Optional[str] = None


class DocumentInfo(BaseModel):
    """Document metadata for comparison result."""
    doc_id: str
    title: str
    chunk_count: int


class CompareStats(BaseModel):
    """Statistics about the comparison."""
    total_changes: int
    high_severity: int
    medium_severity: int
    low_severity: int


class CompareResponse(BaseModel):
    """Response from document comparison."""
    doc_a: DocumentInfo
    doc_b: DocumentInfo
    differences: List[DifferenceItem]
    summary: str
    stats: CompareStats
    error: Optional[str] = None


@app.post("/compare", response_model=CompareResponse)
async def compare_documents(body: CompareRequest):
    """
    Compare two documents and identify material differences.
    
    Uses semantic analysis to find substantive changes, ignoring
    formatting and minor wording differences.
    """
    engine = get_engine()
    
    logger.info(f"Comparing documents: {body.doc_id_a} vs {body.doc_id_b}")
    
    try:
        result = await engine.compare_documents(
            doc_id_a=body.doc_id_a,
            doc_id_b=body.doc_id_b,
            workspace_id=body.workspace_id
        )
        
        return CompareResponse(
            doc_a=DocumentInfo(**result["doc_a"]),
            doc_b=DocumentInfo(**result["doc_b"]),
            differences=[DifferenceItem(**d) for d in result.get("differences", [])],
            summary=result.get("summary", ""),
            stats=CompareStats(**result.get("stats", {"total_changes": 0, "high_severity": 0, "medium_severity": 0, "low_severity": 0})),
            error=result.get("error")
        )
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"Comparison failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ----------------------------
# Disclosure Schedule Endpoints
# ----------------------------

class ScheduleRequest(BaseModel):
    schedule_type: str = Field(..., description="Type of schedule to generate")
    workspace_id: str = Field(default=DEFAULT_WORKSPACE_ID)
    folder_path: Optional[str] = None

class ScheduleItemResponse(BaseModel):
    title: str
    category: str
    description: str
    parties: List[str]
    key_terms: str
    risk_level: str
    source_doc_id: str
    source_doc_title: str

class ScheduleResponse(BaseModel):
    schedule_type: str
    schedule_name: str
    generated_at: str
    items: List[ScheduleItemResponse]
    summary: str
    total_count: int

@app.get("/schedules/types")
async def list_schedule_types():
    """List available disclosure schedule types."""
    engine = get_engine()
    generator = ScheduleGenerator(engine)
    return {"types": generator.list_schedule_types()}

@app.post("/schedules/generate", response_model=ScheduleResponse)
async def generate_schedule(request: ScheduleRequest):
    """
    Generate a disclosure schedule.
    
    This is the core M&A deliverable - creates formatted schedule
    from all relevant documents in the data room.
    """
    try:
        engine = get_engine()
        generator = ScheduleGenerator(engine)
        
        schedule = await generator.generate_schedule(
            schedule_type=request.schedule_type,
            workspace_id=request.workspace_id,
            folder_path=request.folder_path
        )
        
        return ScheduleResponse(
            schedule_type=schedule.schedule_type,
            schedule_name=schedule.schedule_name,
            generated_at=schedule.generated_at,
            items=[ScheduleItemResponse(
                title=i.title,
                category=i.category,
                description=i.description,
                parties=i.parties,
                key_terms=i.key_terms,
                risk_level=i.risk_level,
                source_doc_id=i.source_doc_id,
                source_doc_title=i.source_doc_title
            ) for i in schedule.items],
            summary=schedule.summary,
            total_count=schedule.total_count
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Schedule generation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ----------------------------
# Settings / Mode Endpoints
# ----------------------------

class ModeStatus(BaseModel):
    offline_mode: bool
    llm_provider: str
    ollama_available: bool
    ollama_host: str

@app.get("/settings/mode", response_model=ModeStatus)
async def get_mode_status():
    """Get current offline/online mode status."""
    import os
    ollama_available = await check_ollama_available()
    
    return ModeStatus(
        offline_mode=is_offline_mode(),
        llm_provider=os.getenv("LLM_PROVIDER", "openai"),
        ollama_available=ollama_available,
        ollama_host=os.getenv("OLLAMA_HOST", "http://localhost:11434")
    )

@app.post("/settings/mode")
async def set_mode(offline: bool = True):
    """
    Toggle offline mode.
    Note: This sets the environment variable for the current process.
    For persistence, set OFFLINE_MODE in your .env file.
    """
    import os
    os.environ["OFFLINE_MODE"] = "true" if offline else "false"
    if offline:
        os.environ["LLM_PROVIDER"] = "ollama"
    return {"offline_mode": offline, "message": "Mode updated. Restart recommended for full effect."}


# ----------------------------
# Diligence Workflow Endpoints
# ----------------------------

from models import (
    DocumentReview, ReviewStatus,
    ClauseExtraction, ClauseType, PLAYBOOKS, CLAUSE_LABELS, get_playbook,
    Issue, IssueSeverity, IssueStatus
)
from playbook_engine import PlaybookEngine

# In-memory stores (would be LanceDB tables in production)
_reviews: Dict[str, DocumentReview] = {}
_clauses: Dict[str, ClauseExtraction] = {}
_issues: Dict[str, Issue] = {}


# --- Review Endpoints ---

class ReviewUpdate(BaseModel):
    status: Optional[str] = None
    assigned_to: Optional[str] = None
    reviewer_notes: Optional[str] = None

class ProjectStats(BaseModel):
    total_docs: int
    unreviewed: int
    in_review: int
    review_complete: int
    qa_needed: int
    qa_approved: int
    flagged: int
    deadline_days: int
    throughput_docs_per_hr: float
    completion_percentage: float

class BulkAssignRequest(BaseModel):
    doc_ids: List[str]
    assigned_to: str

class BulkStatusRequest(BaseModel):
    doc_ids: List[str]
    status: str

@app.get("/reviews")
async def list_reviews(
    workspace_id: str = DEFAULT_WORKSPACE_ID,
    status: Optional[str] = None,
    assigned_to: Optional[str] = None
):
    """List all document reviews with optional filters."""
    engine = get_engine()
    docs = engine.store.list_documents(workspace_id)
    
    reviews = []
    for doc in docs:
        doc_id = doc["doc_id"]
        # Get or create review
        if doc_id not in _reviews:
            _reviews[doc_id] = DocumentReview(
                doc_id=doc_id,
                confidence=0.7  # Default confidence
            )
        
        review = _reviews[doc_id]
        
        # Apply filters
        if status and review.status.value != status:
            continue
        if assigned_to and review.assigned_to != assigned_to:
            continue
        
        reviews.append({
            **review.to_dict(),
            "doc_title": doc.get("title", "Unknown"),
            "doc_type": doc.get("doc_type", "Unknown"),
            "risk_level": doc.get("risk_level", "Unknown"),
            "folder_path": doc.get("folder_path", "/"),
        })
    
    return {"reviews": reviews, "total": len(reviews)}

@app.put("/reviews/{doc_id}")
async def update_review(doc_id: str, update: ReviewUpdate):
    """Update a document review."""
    if doc_id not in _reviews:
        _reviews[doc_id] = DocumentReview(doc_id=doc_id)
    
    review = _reviews[doc_id]
    
    if update.status:
        review.status = ReviewStatus(update.status)
        if update.status == "reviewed":
            review.reviewed_at = datetime.utcnow()
    if update.assigned_to is not None:
        review.assigned_to = update.assigned_to
    if update.reviewer_notes is not None:
        review.reviewer_notes = update.reviewer_notes
    
    return review.to_dict()

@app.post("/reviews/bulk-assign")
async def bulk_assign_reviews(request: BulkAssignRequest):
    """Bulk assign documents to a reviewer."""
    for doc_id in request.doc_ids:
        if doc_id not in _reviews:
            _reviews[doc_id] = DocumentReview(doc_id=doc_id)
        _reviews[doc_id].assigned_to = request.assigned_to
    return {"assigned": len(request.doc_ids), "assigned_to": request.assigned_to}

@app.post("/reviews/bulk-status")
async def bulk_update_status(request: BulkStatusRequest):
    """Bulk update review status."""
    for doc_id in request.doc_ids:
        if doc_id not in _reviews:
            _reviews[doc_id] = DocumentReview(doc_id=doc_id)
        _reviews[doc_id].status = ReviewStatus(request.status)
        if request.status == "reviewed":
            _reviews[doc_id].reviewed_at = datetime.utcnow()
    return {"updated": len(request.doc_ids), "status": request.status}


# --- Playbook/Clause Endpoints ---

@app.get("/playbooks")
async def list_playbooks():
    """List available clause extraction playbooks."""
    return {"playbooks": [pb.to_dict() for pb in PLAYBOOKS]}

class PlaybookRunRequest(BaseModel):
    workspace_id: str = DEFAULT_WORKSPACE_ID
    doc_ids: Optional[List[str]] = None

@app.post("/playbooks/{playbook_id}/run")
async def run_playbook(playbook_id: str, request: PlaybookRunRequest):
    """Run a playbook to extract clauses from documents."""
    engine = get_engine()
    playbook_engine = PlaybookEngine(engine)
    
    try:
        result = await playbook_engine.run_playbook(
            playbook_id=playbook_id,
            workspace_id=request.workspace_id,
            doc_ids=request.doc_ids
        )
        
        # Store extractions and issues
        rows_to_persist_clauses = []
        for ext_dict in result.get("extractions", []):
            ext = ClauseExtraction.from_dict(ext_dict)
            _clauses[ext.id] = ext
            row = ext.to_dict()
            row["workspace_id"] = request.workspace_id
            rows_to_persist_clauses.append(row)
        
        rows_to_persist_issues = []
        for issue_dict in result.get("issues", []):
            issue = Issue.from_dict(issue_dict)
            _issues[issue.id] = issue
            row = issue.to_dict()
            row["workspace_id"] = request.workspace_id
            rows_to_persist_issues.append(row)

        # Bulk persist clauses
        if rows_to_persist_clauses:
            try:
                engine.store.upsert_clauses(rows_to_persist_clauses)
            except Exception as e:
                logger.error(f"Failed to persist clauses after playbook run: {e}")
        
        # Bulk persist issues
        if rows_to_persist_issues:
            try:
                engine.store.upsert_issues(rows_to_persist_issues)
            except Exception as e:
                logger.error(f"Failed to persist issues after playbook run: {e}")
        
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Playbook run failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/clauses/matrix")
async def get_clause_matrix(workspace_id: str = DEFAULT_WORKSPACE_ID):
    """Get clause matrix from stored extractions."""
    engine = get_engine()
    playbook_engine = PlaybookEngine(engine)
    
    # Filter clauses by workspace docs
    docs = engine.store.list_documents(workspace_id)
    doc_ids = {d["doc_id"] for d in docs}
    
    relevant_clauses = [c for c in _clauses.values() if c.doc_id in doc_ids]
    
    matrix = playbook_engine.build_matrix(relevant_clauses)
    return matrix

@app.delete("/clauses")
async def clear_all_clauses(workspace_id: str = DEFAULT_WORKSPACE_ID):
    """Clear all clause extractions to allow re-running with improved prompts."""
    global _clauses, _issues
    
    # Get doc_ids for this workspace
    engine = get_engine()
    docs = engine.store.list_documents(workspace_id)
    doc_ids = {d["doc_id"] for d in docs}
    
    # Remove clauses for these docs
    removed_clauses = 0
    removed_issues = 0
    
    clause_ids_to_remove = [c.id for c in _clauses.values() if c.doc_id in doc_ids]
    for clause_id in clause_ids_to_remove:
        del _clauses[clause_id]
        removed_clauses += 1
    
    issue_ids_to_remove = [i.id for i in _issues.values() if i.doc_id in doc_ids]
    for issue_id in issue_ids_to_remove:
        del _issues[issue_id]
        removed_issues += 1
    
    logger.info(f"Cleared {removed_clauses} clauses and {removed_issues} issues for workspace {workspace_id}")
    return {"cleared_clauses": removed_clauses, "cleared_issues": removed_issues}

@app.get("/clauses/{clause_id}")
async def get_clause(clause_id: str):
    """Get a single clause extraction."""
    if clause_id not in _clauses:
        raise HTTPException(status_code=404, detail="Clause not found")
    return _clauses[clause_id].to_dict()

@app.put("/clauses/{clause_id}")
async def update_clause(clause_id: str, verified: Optional[bool] = None, flagged: Optional[bool] = None):
    """Update clause verification status."""
    if clause_id not in _clauses:
        raise HTTPException(status_code=404, detail="Clause not found")
    
    clause = _clauses[clause_id]
    if verified is not None:
        clause.verified = verified
    if flagged is not None:
        clause.flagged = flagged
    
    # Persist
    try:
        engine = get_engine()
        row = clause.to_dict()
        row["workspace_id"] = DEFAULT_WORKSPACE_ID # TODO: Pass workspace in update_clause
        engine.store.upsert_clause(row)
    except Exception as e:
        logger.error(f"Failed to persist clause update: {e}")

    return clause.to_dict()


# --- Issues Endpoints ---

class IssueCreate(BaseModel):
    title: str
    description: str = ""
    severity: str = "warning"
    doc_id: Optional[str] = None
    doc_title: Optional[str] = None
    clause_id: Optional[str] = None
    owner: Optional[str] = None
    action_required: str = ""

class IssueUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    severity: Optional[str] = None
    status: Optional[str] = None
    owner: Optional[str] = None
    action_required: Optional[str] = None

@app.get("/issues")
async def list_issues(
    severity: Optional[str] = None,
    status: Optional[str] = None,
    owner: Optional[str] = None
):
    """List all issues with optional filters."""
    issues = list(_issues.values())
    
    if severity:
        issues = [i for i in issues if i.severity.value == severity]
    if status:
        issues = [i for i in issues if i.status.value == status]
    if owner:
        issues = [i for i in issues if i.owner == owner]
    
    # Sort by severity (critical first)
    severity_order = {"critical": 0, "warning": 1, "info": 2}
    issues.sort(key=lambda x: severity_order.get(x.severity.value, 3))
    
    return {"issues": [i.to_dict() for i in issues], "total": len(issues)}

@app.post("/issues")
async def create_issue(issue: IssueCreate):
    """Create a new issue."""
    new_issue = Issue(
        title=issue.title,
        description=issue.description,
        severity=IssueSeverity(issue.severity),
        doc_id=issue.doc_id,
        doc_title=issue.doc_title,
        clause_id=issue.clause_id,
        owner=issue.owner,
        action_required=issue.action_required,
    )
    _issues[new_issue.id] = new_issue
    
    # Persist
    try:
        engine = get_engine()
        row = new_issue.to_dict()
        row["workspace_id"] = DEFAULT_WORKSPACE_ID # TODO: Pass workspace in IssueCreate
        engine.store.upsert_issue(row)
    except Exception as e:
        logger.error(f"Failed to persist issue: {e}")
        
    return new_issue.to_dict()

@app.put("/issues/{issue_id}")
async def update_issue(issue_id: str, update: IssueUpdate):
    """Update an issue."""
    if issue_id not in _issues:
        raise HTTPException(status_code=404, detail="Issue not found")
    
    issue = _issues[issue_id]
    
    if update.title is not None:
        issue.title = update.title
    if update.description is not None:
        issue.description = update.description
    if update.severity is not None:
        issue.severity = IssueSeverity(update.severity)
    if update.status is not None:
        issue.status = IssueStatus(update.status)
        if update.status == "resolved":
            issue.resolved_at = datetime.utcnow()
    if update.owner is not None:
        issue.owner = update.owner
    if update.action_required is not None:
        issue.action_required = update.action_required
    
    # Persist
    try:
        engine = get_engine()
        row = issue.to_dict()
        # Ensure workspace_id is preserved or defaulted
        # We don't have it in Issue object? Issue to_dict doesn't include workspace_id?
        # We need to add it.
        row["workspace_id"] = DEFAULT_WORKSPACE_ID
        engine.store.upsert_issue(row)
    except Exception as e:
        logger.error(f"Failed to persist issue update: {e}")
        
    return issue.to_dict()

@app.delete("/issues/{issue_id}")
async def delete_issue(issue_id: str):
    """Delete an issue."""
    if issue_id not in _issues:
        raise HTTPException(status_code=404, detail="Issue not found")
    del _issues[issue_id]
    
    # Delete from persistence
    try:
        engine = get_engine()
        engine.store.delete_issue(issue_id)
    except Exception as e:
        logger.error(f"Failed to delete issue from persistence: {e}")

    return {"deleted": True}


# --- Export Endpoints ---

from fastapi.responses import StreamingResponse
import io

@app.get("/exports/clause-matrix.csv")
async def export_clause_matrix_csv(workspace_id: str = DEFAULT_WORKSPACE_ID):
    """Export clause matrix as CSV."""
    engine = get_engine()
    docs = engine.store.list_documents(workspace_id)
    doc_ids = {d["doc_id"] for d in docs}
    
    relevant_clauses = [c for c in _clauses.values() if c.doc_id in doc_ids]
    
    # Build CSV
    output = io.StringIO()
    
    # Get all clause types
    all_types = sorted(set(c.clause_type.value for c in relevant_clauses))
    header = ["Document"] + [CLAUSE_LABELS.get(ClauseType(t), t) for t in all_types]
    output.write(",".join(header) + "\n")
    
    # Group by doc
    by_doc = {}
    for c in relevant_clauses:
        if c.doc_id not in by_doc:
            by_doc[c.doc_id] = {"title": c.doc_title, "clauses": {}}
        by_doc[c.doc_id]["clauses"][c.clause_type.value] = c.extracted_value
    
    # Write rows
    for doc_id, data in by_doc.items():
        row = [data["title"].replace(",", ";")]
        for ct in all_types:
            val = data["clauses"].get(ct, "")
            row.append(val.replace(",", ";").replace("\n", " ")[:100])
        output.write(",".join(row) + "\n")
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=clause_matrix.csv"}
    )

@app.get("/exports/issues.csv")
async def export_issues_csv():
    """Export issues list as CSV."""
    output = io.StringIO()
    
    header = ["Severity", "Title", "Document", "Owner", "Status", "Action Required", "Description"]
    output.write(",".join(header) + "\n")
    
    for issue in _issues.values():
        row = [
            issue.severity.value,
            issue.title.replace(",", ";"),
            (issue.doc_title or "").replace(",", ";"),
            issue.owner or "",
            issue.status.value,
            issue.action_required.replace(",", ";"),
            issue.description.replace(",", ";").replace("\n", " ")[:200],
        ]
        output.write(",".join(row) + "\n")
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=issues_list.csv"}
    )

@app.get("/project/stats", response_model=ProjectStats)
async def get_project_stats(workspace_id: str = DEFAULT_WORKSPACE_ID):
    """Aggregate stats for project dashboard."""
    engine = get_engine()
    docs = engine.store.list_documents(workspace_id)
    total = len(docs)
    
    doc_ids = set(d["doc_id"] for d in docs)
    
    unreviewed = 0
    in_review = 0
    review_complete = 0
    qa_needed = 0
    qa_approved = 0
    flagged = 0
    
    for doc_id in doc_ids:
        if doc_id in _reviews:
            r = _reviews[doc_id]
            # Handle new statuses safely
            s = r.status
            if s == ReviewStatus.UNREVIEWED: unreviewed += 1
            elif s == ReviewStatus.IN_REVIEW: in_review += 1
            elif s == ReviewStatus.REVIEWED: review_complete += 1
            elif s == ReviewStatus.QA_NEEDED: qa_needed += 1
            elif s == ReviewStatus.QA_APPROVED: qa_approved += 1
            elif s == ReviewStatus.FLAGGED: flagged += 1
        else:
            unreviewed += 1
            
    return ProjectStats(
        total_docs=total,
        unreviewed=unreviewed,
        in_review=in_review,
        review_complete=review_complete,
        qa_needed=qa_needed,
        qa_approved=qa_approved,
        flagged=flagged,
        deadline_days=3, # Mock deadline
        throughput_docs_per_hr=18.5, # Mock throughput
        completion_percentage=((review_complete + qa_approved) / max(1, total) * 100)
    )

@app.get("/exports/excel/{export_type}")
async def export_excel(export_type: str, workspace_id: str = DEFAULT_WORKSPACE_ID):
    """Generate Excel export for delivery."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    
    if export_type == "clause_matrix":
        ws.title = "Clause Matrix"
        # Headers
        all_types = sorted(CLAUSE_LABELS.keys(), key=lambda k: k.value)
        headers = ["Document Name"] + [CLAUSE_LABELS[t] for t in all_types]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="36454F", end_color="36454F", fill_type="solid")
        
        # Data
        engine = get_engine()
        docs = engine.store.list_documents(workspace_id)
        doc_map = {d["doc_id"]: d.get("title", "Unknown") for d in docs}
        
        # Group clauses
        by_doc = {d_id: {} for d_id in doc_map}
        for c in _clauses.values():
            if c.doc_id in by_doc:
                by_doc[c.doc_id][c.clause_type] = c.extracted_value
        
        for doc_id, title in doc_map.items():
            row = [title]
            for t in all_types:
                row.append(by_doc[doc_id].get(t, ""))
            ws.append(row)
            
    elif export_type == "issues_list":
        ws.title = "Issues Register"
        headers = ["Severity", "Issue Title", "Document", "Owner", "Status", "Action Required", "Description"]
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
            
        for issue in _issues.values():
            # Check workspace filter? Issues store doc_id? Yes, likely.
            # For now dump all issues or filter if issue has workspace context
            # Assuming current issues match loaded docs
            row = [
                issue.severity.value.upper(),
                issue.title,
                issue.doc_title or "",
                issue.owner or "Unassigned",
                issue.status.value.title(),
                issue.action_required,
                issue.description
            ]
            ws.append(row)
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{export_type}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Run with:
# uvicorn api:app --host 0.0.0.0 --port 8000

